import os
import uuid

import json

import shutil
from datetime import datetime, date, timedelta

from models import (User, AssetRequest, UploadedFile, Approval, ActivityLog, Asset, Bill, 
                   InventoryUpdate, Vendor, ItemAssignment, AssetMaintenance, AssetDepreciation, 
                   WarrantyAlert, ProcurementQuotation, PurchaseOrder)

from dateutil.relativedelta import relativedelta
from flask import render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify
from werkzeug.utils import secure_filename
from sqlalchemy import text, or_, inspect, and_
from app import app, db

@app.template_filter('from_json')
def from_json_filter(value):
    """Custom filter to parse JSON strings in templates"""
    if value:
        try:
            return json.loads(value)
        except:
            return []
    return []
from models import (User, AssetRequest, UploadedFile, Approval, ActivityLog, Asset, Bill, 
                   InventoryUpdate, Vendor, ItemAssignment, AssetMaintenance, AssetDepreciation, 
                   WarrantyAlert, ProcurementQuotation)

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def log_activity(user_id, action, description, request_id=None):
    """Log user activity for audit trail"""
    activity = ActivityLog()
    activity.user_id = user_id
    activity.action = action
    activity.description = description
    activity.request_id = request_id
    activity.ip_address = request.remote_addr
    db.session.add(activity)
    db.session.commit()

def require_login(f):
    """Decorator to require login"""
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

def require_role(roles):
    """Decorator to require specific roles"""
    def decorator(f):
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))

            user = User.query.get(session['user_id'])
            if not user or user.role not in roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        decorated_function.__name__ = f.__name__
        return decorated_function
    return decorator

def can_approve_request(user, asset_request):
    """Check if user can approve/reject a request"""
    # Users cannot approve their own requests
    if user.id == asset_request.user_id:
        return False

    requester = User.query.get(asset_request.user_id)
    if not requester:
        return False

    # MD can approve anything at any level
    if user.role == 'MD':
        return True

    # Approval flow based on requester's role:
    if requester.role in ['User', 'Employee']:  # Regular users
        # User → CM (1) → Admin (2) → SCM (3) → MD (4)
        if user.role == 'Concern Manager' and asset_request.current_approval_level == 1 and user.floor == asset_request.floor:
            return True
        elif user.role == 'Admin' and asset_request.current_approval_level == 2:
            return True
        elif user.role == 'Accounts/SCM' and asset_request.current_approval_level == 3:
            return True
    elif requester.role == 'Concern Manager':
        # CM → Admin (1) → SCM (2) → MD (3)
        if user.role == 'Admin' and asset_request.current_approval_level == 1:
            return True
        elif user.role == 'Accounts/SCM' and asset_request.current_approval_level == 2:
            return True
    elif requester.role == 'Admin':
        # Admin → SCM (1) → MD (2)
        if user.role == 'Accounts/SCM' and asset_request.current_approval_level == 1:
            return True
    elif requester.role == 'Accounts/SCM':
        # SCM → Admin (1) → MD (2)
        if user.role == 'Admin' and asset_request.current_approval_level == 1:
            return True

    return False

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password) and user.is_active:
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            session['full_name'] = user.full_name

            log_activity(user.id, 'Login', f'User {username} logged in successfully')
            flash(f'Welcome back, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')

    return render_template('login.html')

@app.route('/logout')
@require_login
def logout():
    user_id = session.get('user_id')
    username = session.get('username')

    if user_id:
        log_activity(user_id, 'Logout', f'User {username} logged out')

    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@require_login
def dashboard():
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        flash('User not found. Please log in again.', 'danger')
        return redirect(url_for('login'))

    # Floor-specific filtering
    if user.role in ['Admin', 'Accounts/SCM', 'MD']:
        # These roles can see all floors
        user_requests = AssetRequest.query.filter_by(user_id=user.id).order_by(AssetRequest.created_at.desc()).limit(5).all()
    else:
        # Regular users and concern managers see only their floor
        user_requests = AssetRequest.query.filter_by(user_id=user.id, floor=user.floor).order_by(AssetRequest.created_at.desc()).limit(5).all()

    pending_requests = []
    if user.role == 'Concern Manager':
        # CM sees regular user requests at level 1 on their floor
        pending_requests = AssetRequest.query.join(User, AssetRequest.user_id == User.id).filter(
            AssetRequest.status == 'Pending', 
            AssetRequest.current_approval_level == 1, 
            AssetRequest.floor == user.floor,
            User.role.in_(['User', 'Employee'])
        ).limit(5).all()
    elif user.role == 'Admin':
        # Admin sees:
        # - Regular user requests at level 2
        # - CM requests at level 1
        # - SCM requests at level 1
        pending_requests = AssetRequest.query.join(User, AssetRequest.user_id == User.id).filter(
            AssetRequest.status == 'Pending',
            or_(
                and_(AssetRequest.current_approval_level == 2, User.role.in_(['User', 'Employee'])),
                and_(AssetRequest.current_approval_level == 1, User.role == 'Concern Manager'),
                and_(AssetRequest.current_approval_level == 1, User.role == 'Accounts/SCM')
            )
        ).limit(5).all()
    elif user.role == 'Accounts/SCM':
        # SCM sees:
        # - Regular user requests at level 3
        # - CM requests at level 2
        # - Admin requests at level 1
        pending_requests = AssetRequest.query.join(User, AssetRequest.user_id == User.id).filter(
            AssetRequest.status == 'Pending',
            or_(
                and_(AssetRequest.current_approval_level == 3, User.role.in_(['User', 'Employee'])),
                and_(AssetRequest.current_approval_level == 2, User.role == 'Concern Manager'),
                and_(AssetRequest.current_approval_level == 1, User.role == 'Admin')
            )
        ).limit(5).all()
    elif user.role == 'MD':
        # MD sees all pending requests (can approve anything)
        pending_requests = AssetRequest.query.filter(
            AssetRequest.status == 'Pending'
        ).limit(5).all()

    # Get low stock alerts for consumable assets
    low_stock_assets = []
    if user.role in ['Admin', 'MD', 'Accounts/SCM']:
        low_stock_assets = Asset.query.filter(
            Asset.asset_type == 'Consumable Asset',
            Asset.current_quantity <= Asset.minimum_threshold
        ).all()

    recent_activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(10).all()

    # Floor-specific stats
    if user.role in ['Admin', 'Accounts/SCM', 'MD']:
        # These roles see all floors' stats
        stats = {
            'total_requests': AssetRequest.query.count(),
            'pending_requests': AssetRequest.query.filter_by(status='Pending').count(),
            'approved_requests': AssetRequest.query.filter_by(status='Approved').count(),
            'rejected_requests': AssetRequest.query.filter_by(status='Rejected').count(),
            'fulfilled_requests': AssetRequest.query.filter_by(status='Fulfilled').count(),
            'my_requests': AssetRequest.query.filter_by(user_id=user.id).count(),
            'total_assets': Asset.query.count(),
            'available_assets': Asset.query.filter_by(status='Available').count(),
            'in_use_assets': Asset.query.filter_by(status='In Use').count(),
            'maintenance_assets': Asset.query.filter_by(status='Maintenance').count(),
            'total_vendors': Vendor.query.filter_by(is_active=True).count(),
            'total_assignments': ItemAssignment.query.count(),
            'pending_assignments': ItemAssignment.query.filter_by(delivery_status='Pending').count(),
            'delivered_assignments': ItemAssignment.query.filter_by(delivery_status='Delivered').count()
        }
    else:
        # Floor-specific stats for users and concern managers
        stats = {
            'total_requests': AssetRequest.query.filter_by(floor=user.floor).count(),
            'pending_requests': AssetRequest.query.filter_by(status='Pending', floor=user.floor).count(),
            'approved_requests': AssetRequest.query.filter_by(status='Approved', floor=user.floor).count(),
            'rejected_requests': AssetRequest.query.filter_by(status='Rejected', floor=user.floor).count(),
            'fulfilled_requests': AssetRequest.query.filter_by(status='Fulfilled', floor=user.floor).count(),
            'my_requests': AssetRequest.query.filter_by(user_id=user.id).count(),
            'total_assets': Asset.query.count(),
            'available_assets': Asset.query.filter_by(status='Available').count(),
            'in_use_assets': Asset.query.filter_by(status='In Use').count(),
            'maintenance_assets': Asset.query.filter_by(status='Maintenance').count(),
            'total_vendors': Vendor.query.filter_by(is_active=True).count(),
            'total_assignments': ItemAssignment.query.count(),
            'pending_assignments': ItemAssignment.query.filter_by(delivery_status='Pending').count(),
            'delivered_assignments': ItemAssignment.query.filter_by(delivery_status='Delivered').count()
        }

    return render_template('dashboard.html', 
                         user=user, 
                         user_requests=user_requests,
                         pending_requests=pending_requests,
                         recent_activities=recent_activities,
                         stats=stats,
                         low_stock_assets=low_stock_assets)

@app.route('/request', methods=['GET', 'POST'])
@require_login
def create_request():
    if request.method == 'POST':
        asset_request = AssetRequest()
        
        entry_type = request.form.get('entry_type', 'single')
        
        if entry_type == 'bulk':
            # Handle bulk request
            asset_request.is_bulk_request = True
            
            bulk_item_names = request.form.getlist('bulk_item_name[]')
            bulk_quantities = request.form.getlist('bulk_quantity[]')
            bulk_costs = request.form.getlist('bulk_cost[]')
            
            # Filter out empty rows
            bulk_items = []
            total_cost = 0
            for i, name in enumerate(bulk_item_names):
                if name.strip():
                    quantity = int(bulk_quantities[i]) if bulk_quantities[i] else 1
                    cost = float(bulk_costs[i]) if bulk_costs[i] else 0
                    bulk_items.append({
                        'name': name.strip(),
                        'quantity': quantity,
                        'estimated_cost': cost
                    })
                    total_cost += cost * quantity
            
            if not bulk_items:
                flash('Please add at least one item for bulk request.', 'danger')
                return render_template('request_form.html')
            
            asset_request.bulk_items = json.dumps(bulk_items)
            asset_request.item_name = f"Bulk Request - {len(bulk_items)} items"
            asset_request.quantity = sum(item['quantity'] for item in bulk_items)
            asset_request.estimated_cost = total_cost if total_cost > 0 else None
        else:
            # Handle single request
            asset_request.is_bulk_request = False
            asset_request.item_name = request.form['item_name']
            asset_request.quantity = int(request.form['quantity'])
            asset_request.estimated_cost = float(request.form['estimated_cost']) if request.form['estimated_cost'] else None
        
        asset_request.purpose = request.form['purpose']
        asset_request.request_type = request.form['request_type']
        asset_request.urgency = request.form['urgency']
        asset_request.user_id = session['user_id']
        
        # Set floor from the requesting user
        requesting_user = User.query.get(session['user_id'])
        asset_request.floor = requesting_user.floor

        db.session.add(asset_request)
        db.session.flush()   

        uploaded_files = request.files.getlist('files')
        for file in uploaded_files:
            if file and file.filename and allowed_file(file.filename):
                filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)

                uploaded_file = UploadedFile()
                uploaded_file.filename = filename
                uploaded_file.original_filename = file.filename
                uploaded_file.file_path = file_path
                uploaded_file.file_size = os.path.getsize(file_path)
                uploaded_file.mime_type = file.content_type
                uploaded_file.request_id = asset_request.id
                db.session.add(uploaded_file)

        db.session.commit()

        request_type_desc = "bulk" if entry_type == 'bulk' else "single item"
        log_activity(session['user_id'], 'Request Created', 
                    f'Created new {request_type_desc} {asset_request.request_type} request: {asset_request.item_name}',
                    asset_request.id)

        flash('Your request has been submitted successfully!', 'success')
        return redirect(url_for('view_requests'))

    return render_template('request_form.html')

@app.route('/requests')
@require_login
def view_requests():
    user = User.query.get(session['user_id'])
    page = request.args.get('page', 1, type=int)

    if user.role in ['Admin', 'MD', 'Accounts/SCM']:
        # These roles can see all requests
        requests = AssetRequest.query.order_by(AssetRequest.created_at.desc()).paginate(
            page=page, per_page=10, error_out=False)
    elif user.role == 'Concern Manager':
        # Concern managers see only their floor's requests
        requests = AssetRequest.query.filter_by(floor=user.floor).order_by(AssetRequest.created_at.desc()).paginate(
            page=page, per_page=10, error_out=False)
    else:
        # Regular users see only their own requests
        requests = AssetRequest.query.filter_by(user_id=user.id).order_by(AssetRequest.created_at.desc()).paginate(
            page=page, per_page=10, error_out=False)

    # Get assets for availability checking (for SCM role)
    assets = Asset.query.all() if user.role == 'Accounts/SCM' else []
    return render_template('requests.html', requests=requests, user=user, assets=assets)

@app.route('/approve/<int:request_id>/<action>')
@require_login
def approve_request(request_id, action):
    user = User.query.get(session['user_id'])
    asset_request = AssetRequest.query.get_or_404(request_id)

    if not can_approve_request(user, asset_request):
        flash('You do not have permission to approve this request at this level.', 'danger')
        return redirect(url_for('view_requests'))

    approval = Approval()
    approval.request_id = request_id
    approval.approver_id = user.id
    approval.approval_level = asset_request.current_approval_level
    approval.action = action
    approval.comments = request.args.get('comments', '')
    db.session.add(approval)

    if action == 'Approved':
        requester = User.query.get(asset_request.user_id)

        # MD can approve at any level - final approval
        if user.role == 'MD':
            asset_request.status = 'Approved'
        else:
            # Determine max levels based on requester role
            max_levels = 4  # Default for regular users
            if requester.role == 'Concern Manager':
                max_levels = 3  # CM → Admin → SCM → MD
            elif requester.role == 'Admin':
                max_levels = 2  # Admin → SCM → MD
            elif requester.role == 'Accounts/SCM':
                max_levels = 2  # SCM → Admin → MD

            # Check if we've reached the final level
            if asset_request.current_approval_level >= max_levels:
                asset_request.status = 'Approved'
            else:
                asset_request.current_approval_level += 1
                asset_request.status = 'Pending'
    elif action == 'Rejected':
        asset_request.status = 'Rejected'

    asset_request.updated_at = datetime.utcnow()
    db.session.commit()

    log_activity(user.id, f'Request {action}', 
                f'{action} request #{request_id} - {asset_request.item_name}',
                request_id)

    # If SCM approves and matching assets are available, auto-fulfill from assets
    if (action == 'Approved' and user.role == 'Accounts/SCM'):
        matching_assets = Asset.query.filter(
            Asset.name.ilike(f'%{asset_request.item_name}%')
        ).filter(Asset.status.in_(['Available', 'In Use'])).all()

        # Find the best matching asset for auto-fulfillment
        best_asset = None
        for asset in matching_assets:
            if asset.asset_type == 'Consumable Asset':
                if asset.current_quantity >= asset_request.quantity:
                    best_asset = asset
                    break
            elif asset.asset_type == 'Fixed Asset' and asset.status == 'Available':
                best_asset = asset
                break

        if best_asset:
            inventory_update = None  # Initialize the variable

            # Auto-fulfill the request from the matching asset
            if best_asset.asset_type == 'Consumable Asset':
                # Update asset quantity
                previous_quantity = best_asset.current_quantity
                best_asset.current_quantity -= asset_request.quantity
                best_asset.updated_at = datetime.utcnow()

                # Create inventory update record
                inventory_update = InventoryUpdate()
                inventory_update.asset_id = best_asset.id
                inventory_update.previous_quantity = previous_quantity
                inventory_update.new_quantity = best_asset.current_quantity
                inventory_update.update_type = 'Consumption'
                inventory_update.reason = f'Auto-fulfilled request #{asset_request.id} for {asset_request.requester.full_name}'
                inventory_update.updated_by = user.id
                db.session.add(inventory_update)

            elif best_asset.asset_type == 'Fixed Asset':
                # Mark asset as assigned/in use
                best_asset.status = 'In Use'
                best_asset.assigned_to = asset_request.user_id
                best_asset.updated_at = datetime.utcnow()

            # For consumable assets, if quantity becomes 0, mark as out of stock
            if best_asset.asset_type == 'Consumable Asset' and best_asset.current_quantity == 0:
                best_asset.status = 'Out of Stock'

            # Update request fulfillment details
            asset_request.fulfilled_from_asset_id = best_asset.id
            asset_request.fulfilled_quantity = asset_request.quantity
            asset_request.fulfilled_by = user.id
            asset_request.fulfilled_at = datetime.utcnow()
            asset_request.fulfillment_notes = f'Auto-fulfilled by SCM from asset {best_asset.asset_tag}'
            asset_request.status = 'Fulfilled'

            # Show appropriate message based on asset type
            if best_asset.asset_type == 'Consumable Asset' and inventory_update:
                flash(f'Request approved and fulfilled from asset {best_asset.asset_tag}! Asset quantity reduced from {inventory_update.previous_quantity} to {best_asset.current_quantity}.', 'success')
            else:
                flash(f'Request approved and fulfilled from asset {best_asset.asset_tag}!', 'success')
        else:
            # No auto-fulfillment possible, proceed to MD
            flash(f'Request approved by SCM! No matching assets found for auto-fulfillment. Proceeding to MD for final approval.', 'info')
    else:
        if action == 'Approved':
            if user.role == 'MD':
                flash(f'Request has been approved by MD and is now fully approved!', 'success')
            else:
                flash(f'Request has been approved and forwarded to the next level.', 'success')
        else:
            flash(f'Request has been {action.lower()} successfully.', 'success')

    return redirect(url_for('view_requests'))

@app.route('/assign-from-asset/<int:request_id>', methods=['GET', 'POST'])
@require_role(['Accounts/SCM', 'Admin', 'MD'])
def assign_from_asset(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)

    if asset_request.status not in ['Pending', 'Approved']:
        flash('Only pending or approved requests can be assigned from assets.', 'warning')
        return redirect(url_for('view_requests'))

    if request.method == 'POST':
        asset_id = int(request.form['asset_id'])
        assign_quantity = int(request.form['assign_quantity'])
        notes = request.form.get('notes', '')

        asset = Asset.query.get_or_404(asset_id)

        # Validate assignment
        if asset.asset_type == 'Consumable Asset':
            if asset.current_quantity < assign_quantity:
                flash('Insufficient quantity in asset inventory.', 'danger')
                return redirect(url_for('assign_from_asset', request_id=request_id))
        elif asset.asset_type == 'Fixed Asset':
            if assign_quantity > 1:
                flash('Fixed assets can only be assigned with quantity 1.', 'danger')
                return redirect(url_for('assign_from_asset', request_id=request_id))
            if asset.status != 'Available':
                flash('Fixed asset is not available for assignment.', 'danger')
                return redirect(url_for('assign_from_asset', request_id=request_id))

        # Update asset based on type
        if asset.asset_type == 'Consumable Asset':
            previous_quantity = asset.current_quantity
            asset.current_quantity -= assign_quantity
            asset.updated_at = datetime.utcnow()

            # Create inventory update record
            inventory_update = InventoryUpdate()
            inventory_update.asset_id = asset_id
            inventory_update.previous_quantity = previous_quantity
            inventory_update.new_quantity = asset.current_quantity
            inventory_update.update_type = 'Assignment'
            inventory_update.reason = f'Assigned to {asset_request.requester.full_name} for request #{request_id}'
            inventory_update.updated_by = session['user_id']
            db.session.add(inventory_update)

            # Mark as out of stock if quantity reaches 0
            if asset.current_quantity == 0:
                asset.status = 'Out of Stock'

        elif asset.asset_type == 'Fixed Asset':
            asset.status = 'Assigned'
            asset.assigned_to = asset_request.user_id
            asset.updated_at = datetime.utcnow()

        # Update request details
        asset_request.fulfilled_from_asset_id = asset_id
        asset_request.fulfilled_quantity = assign_quantity
        asset_request.fulfilled_by = session['user_id']
        asset_request.fulfilled_at = datetime.utcnow()
        asset_request.fulfillment_notes = notes or f'Assigned from asset {asset.asset_tag}'
        asset_request.status = 'Fulfilled'

        db.session.commit()

        log_activity(session['user_id'], 'Asset Assigned', 
                    f'Assigned asset {asset.asset_tag} to {asset_request.requester.full_name} for request #{request_id}',
                    request_id)

        flash(f'Asset {asset.asset_tag} successfully assigned to {asset_request.requester.full_name}!', 'success')
        return redirect(url_for('view_requests'))

    # Find matching assets for assignment
    matching_assets = Asset.query.filter(
        Asset.name.ilike(f'%{asset_request.item_name}%')
    ).all()

    # Filter available assets
    available_assets = []
    for asset in matching_assets:
        if asset.asset_type == 'Consumable Asset' and asset.current_quantity > 0:
            available_assets.append(asset)
        elif asset.asset_type == 'Fixed Asset' and asset.status == 'Available':
            available_assets.append(asset)

    return render_template('assign_from_asset.html', request=asset_request, assets=available_assets)

@app.route('/admin')
@require_role(['Admin', 'MD'])
def admin_panel():
    users = User.query.all()
    total_users = len(users)
    active_users = len([u for u in users if u.is_active])

    total_requests = AssetRequest.query.count()
    pending_requests = AssetRequest.query.filter_by(status='Pending').count()
    approved_requests = AssetRequest.query.filter_by(status='Approved').count()
    rejected_requests = AssetRequest.query.filter_by(status='Rejected').count()

    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests
    }

    return render_template('admin.html', users=users, stats=stats)

@app.route('/admin/user', methods=['POST'])
@require_role(['Admin', 'MD'])
def create_user():
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    role = request.form['role']
    full_name = request.form['full_name']
    floor = request.form.get('floor', '')
    department = request.form.get('department', '')

    if User.query.filter_by(username=username).first():
        flash('Username already exists.', 'danger')
        return redirect(url_for('admin_panel'))

    if User.query.filter_by(email=email).first():
        flash('Email already exists.', 'danger')
        return redirect(url_for('admin_panel'))

    user = User()
    user.username = username
    user.email = email
    user.role = role
    user.full_name = full_name
    user.floor = floor
    user.department = department
    user.set_password(password)

    db.session.add(user)
    db.session.commit()

    log_activity(session['user_id'], 'User Created', f'Created new user: {username} with role: {role}')
    flash(f'User {username} created successfully.', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/user/<int:user_id>/toggle')
@require_role(['Admin', 'MD'])
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_active = not user.is_active
    db.session.commit()

    status = 'activated' if user.is_active else 'deactivated'
    log_activity(session['user_id'], 'User Status Changed', f'User {user.username} {status}')
    flash(f'User {user.username} has been {status}.', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/vendor/<int:vendor_id>/toggle')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def toggle_vendor(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.is_active = not vendor.is_active
    db.session.commit()

    status = 'activated' if vendor.is_active else 'deactivated'
    log_activity(session['user_id'], 'Vendor Status Changed', f'Vendor {vendor.vendor_name} {status}')
    flash(f'Vendor {vendor.vendor_name} has been {status}.', 'success')
    return redirect(url_for('view_vendors'))

@app.route('/activity')
@require_login
def activity_log():
    user = User.query.get(session['user_id'])
    page = request.args.get('page', 1, type=int)
    activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).paginate(
        page=page, per_page=20, error_out=False)

    return render_template('activity.html', activities=activities, user=user)

@app.route('/uploads/<filename>')
@require_login
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.errorhandler(404)
def not_found(error):
    return render_template('base.html', error='Page not found'), 404



@app.route('/assets')
@require_login
def view_assets():
    page = request.args.get('page', 1, type=int)
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    asset_type = request.args.get('asset_type', '')

    query = Asset.query
    if category:
        query = query.filter_by(category=category)
    if asset_type:
        query = query.filter_by(asset_type=asset_type)
    if status == 'low_stock':
        query = query.filter(
            Asset.asset_type == 'Consumable Asset',
            Asset.current_quantity <= Asset.minimum_threshold
        )
    elif status:
        query = query.filter_by(status=status)

    assets = query.order_by(Asset.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    categories = db.session.query(Asset.category).distinct().all()
    categories = [c[0] for c in categories]

    user = User.query.get(session['user_id'])
    return render_template('assets.html', assets=assets, categories=categories, 
                         selected_category=category, selected_status=status, 
                         selected_asset_type=asset_type, user=user)

@app.route('/asset/add', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def add_asset():
    if request.method == 'POST':
        asset = Asset()
        asset.asset_tag = request.form['asset_tag']
        asset.name = request.form['name']
        asset.category = request.form['category']
        asset.asset_type = request.form['asset_type']
        asset.brand = request.form.get('brand', '')
        asset.model = request.form.get('model', '')
        asset.serial_number = request.form.get('serial_number', '')
        asset.purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d').date() if request.form['purchase_date'] else None
        asset.purchase_cost = float(request.form['purchase_cost']) if request.form['purchase_cost'] else None
        asset.current_value = float(request.form['current_value']) if request.form['current_value'] else None
        asset.condition = request.form['condition']
        asset.location = request.form.get('location', '')
        asset.assigned_to = int(request.form['assigned_to']) if request.form['assigned_to'] else None
        asset.status = request.form['status']
        asset.warranty_expiry = datetime.strptime(request.form['warranty_expiry'], '%Y-%m-%d').date() if request.form['warranty_expiry'] else None
        asset.notes = request.form.get('notes', '')

        # Set inventory fields for consumable assets
        if asset.asset_type == 'Consumable Asset':
            asset.current_quantity = int(request.form.get('current_quantity', 1))
            asset.minimum_threshold = int(request.form.get('minimum_threshold', 5))
            asset.unit_of_measurement = request.form.get('unit_of_measurement', 'Piece')

        try:
            db.session.add(asset)
            db.session.commit()

            log_activity(session['user_id'], 'Asset Added', f'Added new asset: {asset.asset_tag} - {asset.name}')
            flash('Asset added successfully!', 'success')
            return redirect(url_for('view_assets'))
        except Exception as e:
            db.session.rollback()
            if 'duplicate key value violates unique constraint' in str(e) and 'asset_tag' in str(e):
                flash(f'Asset tag "{asset.asset_tag}" already exists. Please use a unique asset tag.', 'danger')
            else:
                flash('Error adding asset. Please try again.', 'danger')
            # Return to form with current data
            users = User.query.filter_by(is_active=True).all()
            return render_template('add_asset.html', users=users)

    users = User.query.filter_by(is_active=True).all()
    return render_template('add_asset.html', users=users)

@app.route('/asset/bulk-upload', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def bulk_upload_assets():
    if request.method == 'POST':
        uploaded_file = request.files.get('bulk_file')
        if not uploaded_file or not uploaded_file.filename:
            flash('Please select a file to upload.', 'danger')
            return redirect(url_for('bulk_upload_assets'))

        filename = secure_filename(uploaded_file.filename)
        file_extension = filename.rsplit('.', 1)[1].lower()

        if file_extension not in ['csv', 'xlsx', 'xls']:
            flash('Please upload a CSV or Excel file.', 'danger')
            return redirect(url_for('bulk_upload_assets'))

        try:
            import pandas as pd
            import io

            # Read the file based on extension
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    # Check if asset_tag already exists
                    existing_asset = Asset.query.filter_by(asset_tag=str(row.get('asset_tag', '')).strip()).first()
                    if existing_asset:
                        errors.append(f"Row {index + 2}: Asset tag '{row.get('asset_tag')}' already exists")
                        error_count += 1
                        continue

                    asset = Asset()
                    asset.asset_tag = str(row.get('asset_tag', '')).strip()
                    asset.name = str(row.get('name', '')).strip()
                    asset.category = str(row.get('category', 'Other')).strip()
                    asset.asset_type = str(row.get('asset_type', 'Fixed Asset')).strip()
                    asset.brand = str(row.get('brand', '')).strip() if pd.notna(row.get('brand')) else ''
                    asset.model = str(row.get('model', '')).strip() if pd.notna(row.get('model')) else ''
                    asset.serial_number = str(row.get('serial_number', '')).strip() if pd.notna(row.get('serial_number')) else ''
                    asset.condition = str(row.get('condition', 'Good')).strip()
                    asset.location = str(row.get('location', '')).strip() if pd.notna(row.get('location')) else ''
                    asset.status = str(row.get('status', 'Available')).strip()
                    asset.notes = str(row.get('notes', '')).strip() if pd.notna(row.get('notes')) else ''

                    # Handle numeric fields
                    if pd.notna(row.get('purchase_cost')):
                        asset.purchase_cost = float(row.get('purchase_cost', 0))
                    if pd.notna(row.get('current_value')):
                        asset.current_value = float(row.get('current_value', 0))

                    # Handle date fields
                    if pd.notna(row.get('purchase_date')):
                        try:
                            asset.purchase_date = pd.to_datetime(row.get('purchase_date')).date()
                        except:
                            pass

                    if pd.notna(row.get('warranty_expiry')):
                        try:
                            asset.warranty_expiry = pd.to_datetime(row.get('warranty_expiry')).date()
                        except:
                            pass

                    # Handle inventory fields for consumable assets
                    if asset.asset_type == 'Consumable Asset':
                        asset.current_quantity = int(row.get('current_quantity', 1))
                        asset.minimum_threshold = int(row.get('minimum_threshold', 5))
                        asset.unit_of_measurement = str(row.get('unit_of_measurement', 'Piece')).strip()

                    # Validate required fields
                    if not asset.asset_tag or not asset.name:
                        errors.append(f"Row {index + 2}: Asset tag and name are required")
                        error_count += 1
                        continue

                    db.session.add(asset)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1

            if success_count > 0:
                db.session.commit()
                log_activity(session['user_id'], 'Bulk Assets Upload', 
                            f'Successfully uploaded {success_count} assets from {filename}')

            if error_count > 0:
                flash(f'Upload completed with {success_count} successful and {error_count} failed entries. Errors: {"; ".join(errors[:5])}', 'warning')
            else:
                flash(f'Successfully uploaded {success_count} assets!', 'success')

        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')

        return redirect(url_for('view_assets'))

    return render_template('bulk_upload_assets.html')

@app.route('/vendor/bulk-upload', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def bulk_upload_vendors():
    if request.method == 'POST':
        uploaded_file = request.files.get('bulk_file')
        if not uploaded_file or not uploaded_file.filename:
            flash('Please select a file to upload.', 'danger')
            return redirect(url_for('bulk_upload_vendors'))

        filename = secure_filename(uploaded_file.filename)
        file_extension = filename.rsplit('.', 1)[1].lower()

        if file_extension not in ['csv', 'xlsx', 'xls']:
            flash('Please upload a CSV or Excel file.', 'danger')
            return redirect(url_for('bulk_upload_vendors'))

        try:
            import pandas as pd
            import io

            # Read the file based on extension
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    # Check if vendor_code already exists (if provided)
                    vendor_code = str(row.get('vendor_code', '')).strip()
                    if vendor_code:
                        existing_vendor = Vendor.query.filter_by(vendor_code=vendor_code).first()
                        if existing_vendor:
                            errors.append(f"Row {index + 2}: Vendor code '{vendor_code}' already exists")
                            error_count += 1
                            continue

                    vendor = Vendor()
                    vendor.vendor_name = str(row.get('vendor_name', '')).strip()
                    vendor.vendor_code = vendor_code
                    vendor.category = str(row.get('category', '')).strip() if pd.notna(row.get('category')) else ''
                    vendor.contact_person = str(row.get('contact_person', '')).strip() if pd.notna(row.get('contact_person')) else ''
                    vendor.phone = str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else ''
                    vendor.email = str(row.get('email', '')).strip() if pd.notna(row.get('email')) else ''
                    vendor.address = str(row.get('address', '')).strip() if pd.notna(row.get('address')) else ''
                    vendor.payment_terms = str(row.get('payment_terms', '')).strip() if pd.notna(row.get('payment_terms')) else ''
                    vendor.notes = str(row.get('notes', '')).strip() if pd.notna(row.get('notes')) else ''
                    vendor.is_active = True

                    # Validate required fields
                    if not vendor.vendor_name:
                        errors.append(f"Row {index + 2}: Vendor name is required")
                        error_count += 1
                        continue

                    db.session.add(vendor)
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1

            if success_count > 0:
                db.session.commit()
                log_activity(session['user_id'], 'Bulk Vendors Upload', 
                            f'Successfully uploaded {success_count} vendors from {filename}')

            if error_count > 0:
                flash(f'Upload completed with {success_count} successful and {error_count} failed entries. Errors: {"; ".join(errors[:5])}', 'warning')
            else:
                flash(f'Successfully uploaded {success_count} vendors!', 'success')

        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'danger')

        return redirect(url_for('view_vendors'))

    return render_template('bulk_upload_vendors.html')

@app.route('/asset/<int:asset_id>')
@require_login
def view_asset_detail(asset_id):
    from datetime import date
    asset = Asset.query.get_or_404(asset_id)
    # Get fulfilled requests for this asset
    fulfilled_requests = AssetRequest.query.filter_by(fulfilled_from_asset_id=asset_id).order_by(AssetRequest.fulfilled_at.desc()).all()
    return render_template('asset_detail.html', asset=asset, today=date.today(), fulfilled_requests=fulfilled_requests)

@app.route('/download/<filename>')
@require_login
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

@app.route('/bills')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def view_bills():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')

    query = Bill.query
    if status:
        query = query.filter_by(status=status)

    bills = query.order_by(Bill.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    user = User.query.get(session['user_id'])
    return render_template('bills.html', bills=bills, selected_status=status, user=user)

@app.route('/bill/upload/<int:request_id>', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def upload_bill(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)

    if request.method == 'POST':
        bill = Bill()
        bill.bill_number = request.form['bill_number']
        bill.vendor_name = request.form['vendor_name']
        bill.bill_amount = float(request.form['bill_amount'])
        try:
            bill.bill_date = datetime.strptime(request.form['bill_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid bill date format.', 'danger')
            return render_template('upload_bill.html', request=asset_request)
        bill.description = request.form.get('description', '')
        bill.request_id = request_id
        bill.uploaded_by = session['user_id']

        # Handle file upload
        bill_file = request.files.get('bill_file')
        if bill_file and bill_file.filename and allowed_file(bill_file.filename):
            filename = str(uuid.uuid4()) + '_' + secure_filename(bill_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            bill_file.save(file_path)
            bill.bill_file_path = file_path
            bill.bill_filename = filename

        db.session.add(bill)
        db.session.commit()

        log_activity(session['user_id'], 'Bill Uploaded', 
                    f'Uploaded bill {bill.bill_number} for request #{request_id}', request_id)

        flash('Bill uploaded successfully!', 'success')
        return redirect(url_for('view_bills'))

    return render_template('upload_bill.html', request=asset_request)

@app.route('/scm/upload-bill', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def scm_upload_bill():
    if request.method == 'POST':
        # Get approved requests that don't have bills yet
        approved_requests = AssetRequest.query.filter_by(status='Approved').all()
        approved_request_ids = [r.id for r in approved_requests]
        existing_bill_request_ids = [b.request_id for b in Bill.query.all()]
        available_requests = [r for r in approved_requests if r.id not in existing_bill_request_ids]

        request_id = int(request.form['request_id'])

        bill = Bill()
        bill.bill_number = request.form['bill_number']
        bill.vendor_name = request.form['vendor_name']
        bill.bill_amount = float(request.form['bill_amount'])
        try:
            bill.bill_date = datetime.strptime(request.form['bill_date'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Invalid bill date format.', 'danger')
            available_requests = []
            return render_template('scm_upload_bill.html', requests=available_requests)
        bill.description = request.form.get('description', '')
        bill.request_id = request_id
        bill.uploaded_by = session['user_id']

        # Handle file upload
        bill_file = request.files.get('bill_file')
        if bill_file and bill_file.filename and allowed_file(bill_file.filename):
            filename = str(uuid.uuid4()) + '_' + secure_filename(bill_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            bill_file.save(file_path)
            bill.bill_file_path = file_path
            bill.bill_filename = filename

        db.session.add(bill)
        db.session.commit()

        log_activity(session['user_id'], 'Bill Uploaded', 
                    f'Uploaded bill {bill.bill_number} for request #{request_id}', request_id)

        flash('Bill uploaded successfully!', 'success')
        return redirect(url_for('view_bills'))

    # Get approved requests that don't have bills yet
    approved_requests = AssetRequest.query.filter_by(status='Approved').all()
    existing_bill_request_ids = [b.request_id for b in Bill.query.all()]
    available_requests = [r for r in approved_requests if r.id not in existing_bill_request_ids]

    return render_template('scm_upload_bill.html', requests=available_requests)

@app.route('/bill/verify/<int:bill_id>/<action>')
@require_role(['Admin', 'MD'])
def verify_bill(bill_id, action):
    bill = Bill.query.get_or_404(bill_id)

    bill.status = 'Verified' if action == 'approve' else 'Rejected'
    bill.verified_by = session['user_id']
    bill.verified_at = datetime.utcnow()
    bill.verification_comments = request.args.get('comments', '')

    db.session.commit()

    log_activity(session['user_id'], f'Bill {action.title()}d', 
                f'{action.title()}d bill {bill.bill_number}')

    flash(f'Bill has been {action}d successfully.', 'success')
    return redirect(url_for('view_bills'))

@app.route('/inventory/update/<int:asset_id>', methods=['GET', 'POST'])
@require_role(['Admin', 'MD'])
def update_inventory(asset_id):
    asset = Asset.query.get_or_404(asset_id)

    if asset.asset_type != 'Consumable Asset':
        flash('Inventory can only be updated for consumable assets.', 'warning')
        return redirect(url_for('view_assets'))

    if request.method == 'POST':
        new_quantity = int(request.form['new_quantity'])
        update_type = request.form['update_type']
        reason = request.form.get('reason', '')

        # Create inventory update record
        inventory_update = InventoryUpdate()
        inventory_update.asset_id = asset_id
        inventory_update.previous_quantity = asset.current_quantity
        inventory_update.new_quantity = new_quantity
        inventory_update.update_type = update_type
        inventory_update.reason = reason
        inventory_update.updated_by = session['user_id']

        # Update asset quantity
        asset.current_quantity = new_quantity
        asset.updated_at = datetime.utcnow()

        db.session.add(inventory_update)
        db.session.commit()

        log_activity(session['user_id'], 'Inventory Updated', 
                    f'Updated inventory for {asset.name} from {inventory_update.previous_quantity} to {new_quantity}')

        flash('Inventory updated successfully!', 'success')
        return redirect(url_for('view_asset_detail', asset_id=asset_id))

    return render_template('update_inventory.html', asset=asset)

@app.route('/edit-request/<int:request_id>', methods=['GET', 'POST'])
@require_role(['Accounts/SCM', 'Admin', 'MD'])
def edit_request(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)

    if asset_request.status not in ['Pending', 'Approved']:
        flash('Only pending or approved requests can be edited.', 'warning')
        return redirect(url_for('view_requests'))

    if request.method == 'POST':
        # Update request details
        asset_request.item_name = request.form['item_name']
        asset_request.quantity = int(request.form['quantity'])
        asset_request.purpose = request.form['purpose']
        asset_request.request_type = request.form['request_type']
        asset_request.estimated_cost = float(request.form['estimated_cost']) if request.form['estimated_cost'] else None
        asset_request.urgency = request.form['urgency']
        asset_request.updated_at = datetime.utcnow()

        db.session.commit()

        log_activity(session['user_id'], 'Request Edited', 
                    f'Edited request #{request_id} - {asset_request.item_name}', request_id)

        flash('Request updated successfully!', 'success')
        return redirect(url_for('view_requests'))

    return render_template('edit_request.html', request=asset_request)

@app.route('/fulfill-request/<int:request_id>', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def fulfill_request(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)

    if asset_request.status != 'Approved':
        flash('Only approved requests can be fulfilled.', 'warning')
        return redirect(url_for('view_requests'))

    if request.method == 'POST':
        asset_id = int(request.form['asset_id'])
        fulfill_quantity = int(request.form['fulfill_quantity'])
        notes = request.form.get('notes', '')

        asset = Asset.query.get_or_404(asset_id)

        # Check if asset has enough quantity
        if asset.asset_type == 'Consumable Asset':
            if asset.current_quantity < fulfill_quantity:
                flash('Insufficient quantity in asset inventory.', 'danger')
                return redirect(url_for('fulfill_request', request_id=request_id))

            # Update asset quantity
            asset.current_quantity -= fulfill_quantity
        elif asset.asset_type == 'Fixed Asset':
            if fulfill_quantity > 1:
                flash('Fixed assets can only be fulfilled with quantity 1.', 'danger')
                return redirect(url_for('fulfill_request', request_id=request_id))

            # Mark asset as assigned/in use
            asset.status = 'In Use'
            asset.assigned_to = asset_request.user_id

        # Update request fulfillment details
        asset_request.fulfilled_from_asset_id = asset_id
        asset_request.fulfilled_quantity = fulfill_quantity
        asset_request.fulfilled_by = session['user_id']
        asset_request.fulfilled_at = datetime.utcnow()
        asset_request.fulfillment_notes = notes
        asset_request.status = 'Fulfilled'

        # Create inventory update record for consumable assets
        if asset.asset_type == 'Consumable Asset':
            inventory_update = InventoryUpdate()
            inventory_update.asset_id = asset_id
            inventory_update.previous_quantity = asset.current_quantity + fulfill_quantity
            inventory_update.new_quantity = asset.current_quantity
            inventory_update.update_type = 'Consumption'
            inventory_update.reason = f'Fulfilled request #{request_id} for {asset_request.requester.full_name}'
            inventory_update.updated_by = session['user_id']
            db.session.add(inventory_update)

        db.session.commit()

        log_activity(session['user_id'], 'Request Fulfilled', 
                    f'Fulfilled request #{request_id} from asset {asset.asset_tag}', request_id)

        flash(f'Request fulfilled successfully from asset {asset.asset_tag}!', 'success')
        return redirect(url_for('view_requests'))

    # Find matching assets
    matching_assets = Asset.query.filter(
        Asset.name.ilike(f'%{asset_request.item_name}%')
    ).filter(Asset.status.in_(['Available', 'In Use'])).all()

    # For consumable assets, only show those with sufficient quantity
    available_assets = []
    for asset in matching_assets:
        if asset.asset_type == 'Consumable Asset':
            if asset.current_quantity >= asset_request.quantity:
                available_assets.append(asset)
        elif asset.asset_type == 'Fixed Asset' and asset.status == 'Available':
            available_assets.append(asset)

    return render_template('fulfill_request.html', request=asset_request, assets=available_assets)

@app.route('/download/requests')
@require_login
def download_requests():
    import csv
    import io
    from flask import make_response

    user = User.query.get(session['user_id'])

    # Get requests based on user role
    if user.role in ['Admin', 'MD']:
        requests = AssetRequest.query.order_by(AssetRequest.created_at.desc()).all()
    else:
        requests = AssetRequest.query.filter_by(user_id=user.id).order_by(AssetRequest.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'ID', 'Item Name', 'Quantity', 'Type', 'Estimated Cost (₹)', 
        'Urgency', 'Status', 'Requester', 'Created Date', 'Purpose'
    ])

    # Write data
    for req in requests:
        writer.writerow([
            req.id, req.item_name, req.quantity, req.request_type,
            req.estimated_cost or 0, req.urgency, req.status,
            req.requester.full_name, req.created_at.strftime('%Y-%m-%d %H:%M'),
            req.purpose[:100] + '...' if len(req.purpose) > 100 else req.purpose
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=requests.csv'

    return response

@app.route('/download/assets')
@require_login
def download_assets():
    import csv
    import io
    from flask import make_response

    assets = Asset.query.order_by(Asset.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Asset Tag', 'Name', 'Category', 'Asset Type', 'Brand', 'Model',
        'Purchase Cost (₹)', 'Current Value (₹)', 'Status', 'Location',
        'Current Quantity', 'Unit', 'Created Date'
    ])

    # Write data
    for asset in assets:
        writer.writerow([
            asset.asset_tag, asset.name, asset.category, asset.asset_type,
            asset.brand or '', asset.model or '', asset.purchase_cost or 0,
            asset.current_value or 0, asset.status, asset.location or '',
            asset.current_quantity, asset.unit_of_measurement,
            asset.created_at.strftime('%Y-%m-%d %H:%M') if asset.created_at else ''
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=assets.csv'

    return response

@app.route('/download/bills')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def download_bills():
    import csv
    import io
    from flask import make_response

    bills = Bill.query.order_by(Bill.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Bill Number', 'Vendor Name', 'Amount (₹)', 'Bill Date',
        'Status', 'Request ID', 'Uploaded By', 'Created Date'
    ])

    # Write data
    for bill in bills:
        writer.writerow([
            bill.bill_number, bill.vendor_name, bill.bill_amount,
            bill.bill_date.strftime('%Y-%m-%d') if bill.bill_date else '',
            bill.status, bill.request_id, bill.uploader.full_name,
            bill.created_at.strftime('%Y-%m-%d %H:%M') if bill.created_at else ''
        ])

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=bills.csv'

    return response

@app.route('/download/assignments')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def download_assignments():
    format_type = request.args.get('format', 'csv')

    assignments = ItemAssignment.query.order_by(ItemAssignment.created_at.desc()).all()

    if format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from flask import make_response
            import io
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            format_type = 'csv'

    if format_type == 'csv':
        import csv
        import io
        from flask import make_response

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Assignment ID', 'Item Name', 'Quantity', 'Assigned To', 'Assigned By', 
            'Vendor', 'Unit Price (₹)', 'Total Amount (₹)', 'Expected Delivery', 
            'Actual Delivery', 'Delivery Status', 'Notes', 'Created Date'
        ])

        # Write data
        for assignment in assignments:
            writer.writerow([
                assignment.id,
                assignment.item_name,
                assignment.quantity,
                assignment.assignee.full_name,
                assignment.assigner.full_name,
                assignment.vendor.vendor_name,
                assignment.unit_price or 0,
                assignment.total_amount or 0,
                assignment.expected_delivery_date.strftime('%Y-%m-%d') if assignment.expected_delivery_date else '',
                assignment.actual_delivery_date.strftime('%Y-%m-%d') if assignment.actual_delivery_date else '',
                assignment.delivery_status,
                assignment.notes or '',
                assignment.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=item_assignments.csv'
        return response

    # Excel format
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Assignments"

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        'Assignment ID', 'Item Name', 'Quantity', 'Assigned To', 'Department', 
        'Floor', 'Assigned By', 'Vendor', 'Contact Person', 'Unit Price (₹)', 
        'Total Amount (₹)', 'Expected Delivery', 'Actual Delivery', 'Delivery Status', 
        'Notes', 'Delivery Notes', 'Created Date'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Data
    for row, assignment in enumerate(assignments, 2):
        ws.cell(row=row, column=1, value=assignment.id)
        ws.cell(row=row, column=2, value=assignment.item_name)
        ws.cell(row=row, column=3, value=assignment.quantity)
        ws.cell(row=row, column=4, value=assignment.assignee.full_name)
        ws.cell(row=row, column=5, value=assignment.assignee.department)
        ws.cell(row=row, column=6, value=assignment.assignee.floor)
        ws.cell(row=row, column=7, value=assignment.assigner.full_name)
        ws.cell(row=row, column=8, value=assignment.vendor.vendor_name)
        ws.cell(row=row, column=9, value=assignment.vendor.contact_person or '')
        ws.cell(row=row, column=10, value=assignment.unit_price or 0)
        ws.cell(row=row, column=11, value=assignment.total_amount or 0)
        ws.cell(row=row, column=12, value=assignment.expected_delivery_date.strftime('%Y-%m-%d') if assignment.expected_delivery_date else '')
        ws.cell(row=row, column=13, value=assignment.actual_delivery_date.strftime('%Y-%m-%d') if assignment.actual_delivery_date else '')
        ws.cell(row=row, column=14, value=assignment.delivery_status)
        ws.cell(row=row, column=15, value=assignment.notes or '')
        ws.cell(row=row, column=16, value=assignment.delivery_notes or '')
        ws.cell(row=row, column=17, value=assignment.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=item_assignments.xlsx'
    return response

@app.route('/assignment/<int:assignment_id>')
@require_login
def view_assignment_detail(assignment_id):
    assignment = ItemAssignment.query.get_or_404(assignment_id)
    user = User.query.get(session['user_id'])
    return render_template('assignment_detail.html', assignment=assignment, user=user)

@app.route('/download/recent-activity')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def download_recent_activity():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        import io
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        import csv
        import io
        from flask import make_response

        # Get recent activities excluding login/logout
        activities = ActivityLog.query.filter(
            ~ActivityLog.action.in_(['Login', 'Logout'])
        ).order_by(ActivityLog.timestamp.desc()).limit(500).all()

        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([
            'Date', 'Time', 'User', 'Action', 'Description', 'Request ID'
        ])

        # Write data
        for activity in activities:
            writer.writerow([
                activity.timestamp.strftime('%Y-%m-%d'),
                activity.timestamp.strftime('%H:%M:%S'),
                activity.user.full_name,
                activity.action,
                activity.description,
                activity.request_id or ''
            ])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=recent_activity.csv'
        return response

    activities = ActivityLog.query.filter(
        ~ActivityLog.action.in_(['Login', 'Logout'])
    ).order_by(ActivityLog.timestamp.desc()).limit(500).all()

    assignments = ItemAssignment.query.order_by(ItemAssignment.created_at.desc()).limit(200).all()

    requests = AssetRequest.query.order_by(AssetRequest.created_at.desc()).limit(200).all()

    wb = openpyxl.Workbook()

    ws_activity = wb.active
    ws_activity.title = "Recent Activity"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Activity headers
    activity_headers = ['Date', 'Time', 'User', 'Role', 'Action', 'Description', 'Request ID']
    for col, header in enumerate(activity_headers, 1):
        cell = ws_activity.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Activity data
    for row, activity in enumerate(activities, 2):
        ws_activity.cell(row=row, column=1, value=activity.timestamp.strftime('%Y-%m-%d'))
        ws_activity.cell(row=row, column=2, value=activity.timestamp.strftime('%H:%M:%S'))
        ws_activity.cell(row=row, column=3, value=activity.user.full_name)
        ws_activity.cell(row=row, column=4, value=activity.user.role)
        ws_activity.cell(row=row, column=5, value=activity.action)
        ws_activity.cell(row=row, column=6, value=activity.description)
        ws_activity.cell(row=row, column=7, value=activity.request_id or '')

    # Auto-adjust column widths


# SCM Item Classification Route
@app.route('/classify-item/<int:request_id>', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def classify_item(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)
    
    if asset_request.status != 'Approved':
        flash('Only approved requests can be classified.', 'warning')
        return redirect(url_for('view_requests'))
    
    if request.method == 'POST':
        classification = request.form['classification']  # 'Regular' or 'Specific'
        asset_request.item_classification = classification
        
        if classification == 'Regular':
            # For regular items, SCM can directly create PO
            return redirect(url_for('create_purchase_order_from_request', request_id=request_id))
        else:
            # For specific items, redirect to create PO with vendor selection
            return redirect(url_for('create_purchase_order_from_request', request_id=request_id, type='specific'))
    
    return render_template('classify_item.html', request=asset_request)

@app.route('/create-po-from-request/<int:request_id>')
@require_role(['Accounts/SCM'])
def create_purchase_order_from_request(request_id):
    asset_request = AssetRequest.query.get_or_404(request_id)
    item_type = request.args.get('type', 'regular')
    
    if not asset_request.item_classification:
        flash('Please classify the item first.', 'warning')
        return redirect(url_for('classify_item', request_id=request_id))
    
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.vendor_name).all()
    return render_template('create_po_from_request.html', 
                         request=asset_request, 
                         vendors=vendors,
                         item_type=item_type)

# Purchase Order Management Routes
@app.route('/purchase-orders')
@require_role(['Accounts/SCM', 'Admin', 'MD'])
def view_purchase_orders():
    user = User.query.get(session['user_id'])
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    po_type = request.args.get('type', '')

    query = PurchaseOrder.query
    if status:
        query = query.filter_by(status=status)
    if po_type:
        query = query.filter_by(item_type=po_type)

    purchase_orders = query.order_by(PurchaseOrder.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    return render_template('purchase_orders.html', 
                         purchase_orders=purchase_orders,
                         selected_status=status,
                         selected_type=po_type,
                         user=user)

@app.route('/purchase-order/create', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def create_purchase_order():
    if request.method == 'POST':
        po = PurchaseOrder()
        
        # Check if this PO is being created from a request
        source_request_id = request.form.get('source_request_id')
        if source_request_id:
            po.request_id = int(source_request_id)
        
        # Basic Information
        po.item_type = request.form['item_type']
        po.item_name = request.form['item_name']
        po.item_description = request.form.get('item_description', '')
        po.quantity = int(request.form['quantity'])
        po.unit_price = float(request.form['unit_price'])
        po.gst_percentage = float(request.form.get('gst_percentage', 18.0))
        
        # Vendor Information
        po.vendor_id = int(request.form['vendor_id'])
        vendor = Vendor.query.get(po.vendor_id)
        po.vendor_name = vendor.vendor_name
        po.vendor_address = vendor.address
        po.vendor_gst = request.form.get('vendor_gst', '')
        
        # Terms
        po.payment_terms = request.form.get('payment_terms', 'Net 30 days')
        po.delivery_terms = request.form.get('delivery_terms', '')
        po.warranty_terms = request.form.get('warranty_terms', '')
        po.special_instructions = request.form.get('special_instructions', '')
        
        # Expected delivery
        if request.form.get('expected_delivery_date'):
            po.expected_delivery_date = datetime.strptime(request.form['expected_delivery_date'], '%Y-%m-%d').date()
        
        po.created_by = session['user_id']
        
        # Calculate totals
        po.calculate_totals()
        po.generate_po_number()
        
        # Set approval requirements based on item type
        if po.item_type == 'Specific':
            po.requires_md_approval = True
            po.status = 'MD Review Pending'
            
            # Handle file uploads for specific items
            quotation_files = []
            vendor_docs = []
            
            uploaded_quotations = request.files.getlist('quotation_files')
            for file in uploaded_quotations:
                if file and file.filename and allowed_file(file.filename):
                    filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    quotation_files.append(filename)
            
            uploaded_vendor_docs = request.files.getlist('vendor_documents')
            for file in uploaded_vendor_docs:
                if file and file.filename and allowed_file(file.filename):
                    filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(file_path)
                    vendor_docs.append(filename)
            
            po.quotation_files = json.dumps(quotation_files) if quotation_files else None
            po.vendor_documents = json.dumps(vendor_docs) if vendor_docs else None
            
        else:  # Regular item
            po.requires_md_approval = False
            po.md_approved = True
            po.status = 'Approved'
        
        try:
            db.session.add(po)
            db.session.commit()

            log_activity(session['user_id'], 'Purchase Order Created', 
                        f'Created {po.item_type} purchase order {po.po_number} for {po.item_name}')
            
            if po.item_type == 'Regular':
                flash(f'Regular item purchase order {po.po_number} created successfully!', 'success')
            else:
                flash(f'Specific item purchase order {po.po_number} created and sent for MD review!', 'info')
            
            return redirect(url_for('view_purchase_orders'))
            
        except Exception as e:
            db.session.rollback()
            flash('Error creating purchase order. Please try again.', 'danger')
    
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.vendor_name).all()
    requests = AssetRequest.query.filter_by(status='Approved').all()
    return render_template('create_purchase_order.html', vendors=vendors, requests=requests)

@app.route('/purchase-order/<int:po_id>')
@require_role(['Accounts/SCM', 'Admin', 'MD'])
def view_purchase_order_detail(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    user = User.query.get(session['user_id'])
    
    # Parse file attachments
    quotation_files = json.loads(po.quotation_files) if po.quotation_files else []
    vendor_documents = json.loads(po.vendor_documents) if po.vendor_documents else []
    
    return render_template('purchase_order_detail.html', 
                         po=po, 
                         user=user,
                         quotation_files=quotation_files,
                         vendor_documents=vendor_documents)

@app.route('/purchase-order/<int:po_id>/md-review', methods=['GET', 'POST'])
@require_role(['MD'])
def md_review_purchase_order(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if po.status != 'MD Review Pending':
        flash('This purchase order is not pending MD review.', 'warning')
        return redirect(url_for('view_purchase_order_detail', po_id=po_id))
    
    if request.method == 'POST':
        action = request.form['action']  # approve or reject
        po.md_comments = request.form.get('comments', '')
        
        if action == 'approve':
            po.md_approved = True
            po.status = 'Approved'
            po.approved_by_md = session['user_id']
            po.approved_at = datetime.utcnow()
            
            log_activity(session['user_id'], 'PO MD Approved', 
                        f'MD approved purchase order {po.po_number}')
            flash(f'Purchase order {po.po_number} approved! SCM can now generate the final PO.', 'success')
        else:
            po.status = 'MD Rejected'
            
            log_activity(session['user_id'], 'PO MD Rejected', 
                        f'MD rejected purchase order {po.po_number}')
            flash(f'Purchase order {po.po_number} rejected and returned to SCM for updates.', 'info')
        
        po.updated_by = session['user_id']
        db.session.commit()
        
        return redirect(url_for('view_purchase_orders'))
    
    # Parse file attachments for display
    quotation_files = json.loads(po.quotation_files) if po.quotation_files else []
    vendor_documents = json.loads(po.vendor_documents) if po.vendor_documents else []
    
    return render_template('md_review_po.html', 
                         po=po,
                         quotation_files=quotation_files,
                         vendor_documents=vendor_documents)

@app.route('/purchase-order/<int:po_id>/update-vendor', methods=['GET', 'POST'])
@require_role(['Accounts/SCM'])
def update_po_vendor(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if po.status != 'MD Rejected':
        flash('Only MD rejected purchase orders can be updated.', 'warning')
        return redirect(url_for('view_purchase_order_detail', po_id=po_id))
    
    if request.method == 'POST':
        # Update vendor information
        po.vendor_id = int(request.form['vendor_id'])
        vendor = Vendor.query.get(po.vendor_id)
        po.vendor_name = vendor.vendor_name
        po.vendor_address = vendor.address
        po.vendor_gst = request.form.get('vendor_gst', '')
        
        # Update pricing
        po.unit_price = float(request.form['unit_price'])
        po.gst_percentage = float(request.form.get('gst_percentage', 18.0))
        
        # Recalculate totals
        po.calculate_totals()
        
        # Handle new file uploads
        quotation_files = json.loads(po.quotation_files) if po.quotation_files else []
        vendor_docs = json.loads(po.vendor_documents) if po.vendor_documents else []
        
        uploaded_quotations = request.files.getlist('quotation_files')
        for file in uploaded_quotations:
            if file and file.filename and allowed_file(file.filename):
                filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                quotation_files.append(filename)
        
        uploaded_vendor_docs = request.files.getlist('vendor_documents')
        for file in uploaded_vendor_docs:
            if file and file.filename and allowed_file(file.filename):
                filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                vendor_docs.append(filename)
        
        po.quotation_files = json.dumps(quotation_files) if quotation_files else None
        po.vendor_documents = json.dumps(vendor_docs) if vendor_docs else None
        
        # Reset status for MD review
        po.status = 'MD Review Pending'
        po.md_approved = False
        po.md_comments = None
        po.updated_by = session['user_id']
        
        db.session.commit()
        
        log_activity(session['user_id'], 'PO Updated', 
                    f'Updated purchase order {po.po_number} with new vendor and resubmitted for MD review')
        flash(f'Purchase order {po.po_number} updated and resubmitted for MD review!', 'info')
        
        return redirect(url_for('view_purchase_orders'))
    
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.vendor_name).all()
    return render_template('update_po_vendor.html', po=po, vendors=vendors)

@app.route('/purchase-order/<int:po_id>/generate', methods=['POST'])
@require_role(['Accounts/SCM'])
def generate_final_purchase_order(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if po.status != 'Approved':
        flash('Only approved purchase orders can be generated.', 'warning')
        return redirect(url_for('view_purchase_order_detail', po_id=po_id))
    
    po.status = 'Generated'
    po.po_status = 'Created'
    po.updated_by = session['user_id']
    
    db.session.commit()
    
    log_activity(session['user_id'], 'PO Generated', 
                f'Generated final purchase order {po.po_number}')
    flash(f'Purchase order {po.po_number} generated successfully!', 'success')
    
    return redirect(url_for('print_purchase_order', po_id=po_id))

@app.route('/purchase-order/<int:po_id>/print')
@require_role(['Accounts/SCM', 'Admin', 'MD'])
def print_purchase_order(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    return render_template('print_purchase_order.html', po=po)

@app.route('/purchase-order/<int:po_id>/update-status', methods=['POST'])
@require_role(['Accounts/SCM', 'Admin'])
def update_po_status(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    po.po_status = request.form['po_status']
    po.updated_by = session['user_id']
    
    db.session.commit()
    
    log_activity(session['user_id'], 'PO Status Updated', 
                f'Updated purchase order {po.po_number} status to {po.po_status}')
    flash(f'Purchase order status updated to {po.po_status}!', 'success')
    
    return redirect(url_for('view_purchase_order_detail', po_id=po_id))


    for column in ws_activity.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_activity.column_dimensions[column_letter].width = adjusted_width

    # Assignments Sheet
    ws_assignments = wb.create_sheet("Item Assignments")
    assignment_headers = ['ID', 'Item Name', 'Quantity', 'Assigned To', 'Assigned By', 'Vendor', 
                         'Status', 'Amount (₹)', 'Expected Delivery', 'Actual Delivery', 'Created Date']

    for col, header in enumerate(assignment_headers, 1):
        cell = ws_assignments.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row, assignment in enumerate(assignments, 2):
        ws_assignments.cell(row=row, column=1, value=assignment.id)
        ws_assignments.cell(row=row, column=2, value=assignment.item_name)
        ws_assignments.cell(row=row, column=3, value=assignment.quantity)
        ws_assignments.cell(row=row, column=4, value=assignment.assignee.full_name)
        ws_assignments.cell(row=row, column=5, value=assignment.assigner.full_name)
        ws_assignments.cell(row=row, column=6, value=assignment.vendor.vendor_name)
        ws_assignments.cell(row=row, column=7, value=assignment.delivery_status)
        ws_assignments.cell(row=row, column=8, value=assignment.total_amount or 0)
        ws_assignments.cell(row=row, column=9, value=assignment.expected_delivery_date.strftime('%Y-%m-%d') if assignment.expected_delivery_date else '')
        ws_assignments.cell(row=row, column=10, value=assignment.actual_delivery_date.strftime('%Y-%m-%d') if assignment.actual_delivery_date else '')
        ws_assignments.cell(row=row, column=11, value=assignment.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-adjust column widths for assignments
    for column in ws_assignments.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_assignments.column_dimensions[column_letter].width = adjusted_width

    # Requests Sheet
    ws_requests = wb.create_sheet("Recent Requests")
    request_headers = ['ID', 'Item Name', 'Quantity', 'Type', 'Requester', 'Department', 
                      'Status', 'Current Level', 'Estimated Cost (₹)', 'Urgency', 'Created Date']

    for col, header in enumerate(request_headers, 1):
        cell = ws_requests.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    for row, req in enumerate(requests, 2):
        ws_requests.cell(row=row, column=1, value=req.id)
        ws_requests.cell(row=row, column=2, value=req.item_name)
        ws_requests.cell(row=row, column=3, value=req.quantity)
        ws_requests.cell(row=row, column=4, value=req.request_type)
        ws_requests.cell(row=row, column=5, value=req.requester.full_name)
        ws_requests.cell(row=row, column=6, value=req.requester.department)
        ws_requests.cell(row=row, column=7, value=req.status)
        ws_requests.cell(row=row, column=8, value=req.current_approval_level)
        ws_requests.cell(row=row, column=9, value=req.estimated_cost or 0)
        ws_requests.cell(row=row, column=10, value=req.urgency)
        ws_requests.cell(row=row, column=11, value=req.created_at.strftime('%Y-%m-%d %H:%M'))

    # Auto-adjust column widths for requests
    for column in ws_requests.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_requests.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=system_report.xlsx'
    return response

@app.route('/vendors')
@require_login
def view_vendors():
    page = request.args.get('page', 1, type=int)
    category = request.args.get('category', '')

    query = Vendor.query
    if category:
        query = query.filter_by(category=category)

    vendors = query.order_by(Vendor.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    categories = db.session.query(Vendor.category).distinct().all()
    categories = [c[0] for c in categories if c[0]]

    user = User.query.get(session['user_id'])
    return render_template('vendors.html', vendors=vendors, categories=categories,
                         selected_category=category, user=user)

@app.route('/vendor/add', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def add_vendor():
    if request.method == 'POST':
        vendor = Vendor()
        vendor.vendor_name = request.form['vendor_name']
        vendor.vendor_code = request.form.get('vendor_code', '')
        vendor.category = request.form.get('category', '')
        vendor.contact_person = request.form.get('contact_person', '')
        vendor.phone = request.form.get('phone', '')
        vendor.email = request.form.get('email', '')
        vendor.address = request.form.get('address', '')
        vendor.payment_terms = request.form.get('payment_terms', '')
        vendor.notes = request.form.get('notes', '')

        try:
            db.session.add(vendor)
            db.session.commit()

            log_activity(session['user_id'], 'Vendor Added', f'Added new vendor: {vendor.vendor_name}')
            flash('Vendor added successfully!', 'success')
            return redirect(url_for('view_vendors'))
        except Exception as e:
            db.session.rollback()
            if 'duplicate key value violates unique constraint' in str(e) and 'vendor_code' in str(e):
                flash(f'Vendor code "{vendor.vendor_code}" already exists. Please use a unique vendor code.', 'danger')
            else:
                flash('Error adding vendor. Please try again.', 'danger')

    return render_template('add_vendor.html')

@app.route('/vendor/<int:vendor_id>')
@require_login
def view_vendor_detail(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    # Get assignments for this vendor
    assignments = ItemAssignment.query.filter_by(vendor_id=vendor_id).order_by(ItemAssignment.created_at.desc()).limit(10).all()
    user = User.query.get(session['user_id'])
    return render_template('vendor_detail.html', vendor=vendor, assignments=assignments, user=user)

@app.route('/vendor/<int:vendor_id>/edit', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def edit_vendor(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)

    if request.method == 'POST':
        vendor.vendor_name = request.form['vendor_name']
        vendor.vendor_code = request.form.get('vendor_code', '')
        vendor.category = request.form.get('category', '')
        vendor.contact_person = request.form.get('contact_person', '')
        vendor.phone = request.form.get('phone', '')
        vendor.email = request.form.get('email', '')
        vendor.address = request.form.get('address', '')
        vendor.payment_terms = request.form.get('payment_terms', '')
        vendor.notes = request.form.get('notes', '')
        vendor.updated_at = datetime.utcnow()

        try:
            db.session.commit()
            log_activity(session['user_id'], 'Vendor Updated', f'Updated vendor: {vendor.vendor_name}')
            flash('Vendor updated successfully!', 'success')
            return redirect(url_for('view_vendor_detail', vendor_id=vendor_id))
        except Exception as e:
            db.session.rollback()
            flash('Error updating vendor. Please try again.', 'danger')

    return render_template('edit_vendor.html', vendor=vendor)

@app.route('/assignments')
@require_login
def view_assignments():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    vendor_id = request.args.get('vendor_id', type=int)

    query = ItemAssignment.query
    if status:
        query = query.filter_by(delivery_status=status)
    if vendor_id:
        query = query.filter_by(vendor_id=vendor_id)

    assignments = query.order_by(ItemAssignment.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    vendors = Vendor.query.filter_by(is_active=True).all()
    user = User.query.get(session['user_id'])
    return render_template('assignments.html', assignments=assignments, vendors=vendors,
                         selected_status=status, selected_vendor_id=vendor_id, user=user)

@app.route('/assignment/add', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def add_assignment():
    if request.method == 'POST':
        assignment = ItemAssignment()
        assignment.item_name = request.form['item_name']
        assignment.quantity = int(request.form['quantity'])
        assignment.vendor_id = int(request.form['vendor_id'])
        assignment.assigned_by = session['user_id']
        assignment.assigned_to = int(request.form['assigned_to'])
        assignment.expected_delivery_date = datetime.strptime(request.form['expected_delivery_date'], '%Y-%m-%d').date() if request.form['expected_delivery_date'] else None
        assignment.unit_price = float(request.form['unit_price']) if request.form['unit_price'] else None
        assignment.total_amount = assignment.unit_price * assignment.quantity if assignment.unit_price else None
        assignment.notes = request.form.get('notes', '')

        try:
            db.session.add(assignment)
            db.session.commit()

            log_activity(session['user_id'], 'Assignment Created', 
                        f'Assigned {assignment.item_name} to {assignment.assignee.full_name} via vendor {assignment.vendor.vendor_name}')
            flash('Item assignment created successfully!', 'success')
            return redirect(url_for('view_assignments'))
        except Exception as e:
            db.session.rollback()
            flash('Error creating assignment. Please try again.', 'danger')

    vendors = Vendor.query.filter_by(is_active=True).all()
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    return render_template('assign_item.html', vendors=vendors, users=users)

@app.route('/assignment/<int:assignment_id>/deliver', methods=['POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def deliver_assignment(assignment_id):
    assignment = ItemAssignment.query.get_or_404(assignment_id)

    assignment.delivery_status = 'Delivered'
    assignment.actual_delivery_date = datetime.utcnow().date()
    assignment.delivery_notes = request.form.get('delivery_notes', '')

    db.session.commit()

    log_activity(session['user_id'], 'Assignment Delivered', 
                f'Marked assignment #{assignment_id} as delivered')
    flash('Assignment marked as delivered!', 'success')
    return redirect(url_for('view_assignments'))

@app.route('/assignment/<int:assignment_id>/update-status', methods=['POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def update_assignment_status(assignment_id):
    assignment = ItemAssignment.query.get_or_404(assignment_id)

    assignment.delivery_status = request.form['delivery_status']
    assignment.notes = request.form.get('notes', '')
    assignment.updated_at = datetime.utcnow()

    if assignment.delivery_status == 'Delivered' and not assignment.actual_delivery_date:
        assignment.actual_delivery_date = datetime.utcnow().date()

    db.session.commit()

    log_activity(session['user_id'], 'Assignment Status Updated', 
                f'Updated assignment #{assignment_id} status to {assignment.delivery_status}')
    flash('Assignment status updated successfully!', 'success')
    return redirect(url_for('view_assignments'))

@app.route('/asset-assignments')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def view_asset_assignments():
    page = request.args.get('page', 1, type=int)

    # Get all fulfilled requests that were assigned from assets
    fulfilled_requests = AssetRequest.query.filter(
        AssetRequest.status == 'Fulfilled',
        AssetRequest.fulfilled_from_asset_id.isnot(None)
    ).order_by(AssetRequest.fulfilled_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    user = User.query.get(session['user_id'])
    return render_template('asset_assignments.html', assignments=fulfilled_requests, user=user)

# Asset Lifecycle Management Routes
@app.route('/asset-lifecycle')
@require_login
def asset_lifecycle_dashboard():
    user = User.query.get(session['user_id'])

    # Get maintenance schedules
    maintenance_schedules = AssetMaintenance.query.filter_by(
        status='Scheduled'
    ).order_by(AssetMaintenance.scheduled_date).all()

    # Get warranty alerts
    warranty_alerts = WarrantyAlert.query.filter_by(
        alert_type='Warranty Expiry'
    ).order_by(WarrantyAlert.alert_date).all()

    # Get assets with warranty expiring soon
    from datetime import date, timedelta
    today = date.today()
    thirty_days_later = today + timedelta(days=30)
    seven_days_later = today + timedelta(days=7)

    warranty_expiring = Asset.query.filter(
        Asset.warranty_expiry.isnot(None),
        Asset.warranty_expiry <= thirty_days_later,
        Asset.warranty_expiry >= today
    ).order_by(Asset.warranty_expiry).all()

    # Get maintenance due in next 7 days
    maintenance_due = AssetMaintenance.query.filter(
        AssetMaintenance.scheduled_date <= seven_days_later,
        AssetMaintenance.scheduled_date >= today,
        AssetMaintenance.status == 'Scheduled'
    ).order_by(AssetMaintenance.scheduled_date).all()

    # Get active alerts
    active_alerts = WarrantyAlert.query.filter_by(is_active=True).limit(5).all()

    # Calculate statistics
    warranty_expiring_count = len(warranty_expiring)
    maintenance_due_count = len(maintenance_due)
    current_book_value = sum([a.current_value for a in Asset.query.all() if a.current_value]) or 0
    active_alerts_count = WarrantyAlert.query.filter_by(is_active=True).count()

    stats = {
        'warranty_expiring_count': warranty_expiring_count,
        'maintenance_due_count': maintenance_due_count,
        'current_book_value': current_book_value,
        'active_alerts_count': active_alerts_count
    }

    return render_template('asset_lifecycle.html', 
                         user=user,
                         maintenance_schedules=maintenance_schedules,
                         warranty_alerts=warranty_alerts,
                         warranty_expiring=warranty_expiring,
                         maintenance_due=maintenance_due,
                         active_alerts=active_alerts,
                         stats=stats,
                         today=today)

@app.route('/maintenance')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def view_maintenance():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')

    query = AssetMaintenance.query
    if status:
        query = query.filter_by(status=status)

    maintenance_records = query.order_by(AssetMaintenance.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    user = User.query.get(session['user_id'])
    return render_template('maintenance.html', 
                         maintenance_records=maintenance_records,
                         selected_status=status,
                         user=user)

@app.route('/maintenance/add', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def add_maintenance():
    if request.method == 'POST':
        maintenance = AssetMaintenance()
        maintenance.asset_id = int(request.form['asset_id'])
        maintenance.maintenance_type = request.form['maintenance_type']
        maintenance.description = request.form['description']
        maintenance.scheduled_date = datetime.strptime(request.form['scheduled_date'], '%Y-%m-%d').date()
        maintenance.estimated_cost = float(request.form['estimated_cost']) if request.form['estimated_cost'] else None
        maintenance.service_provider = request.form.get('assigned_to', '')
        maintenance.maintenance_notes = request.form.get('notes', '')
        maintenance.created_by = session['user_id']

        db.session.add(maintenance)
        db.session.commit()

        log_activity(session['user_id'], 'Maintenance Scheduled', 
                    f'Scheduled {maintenance.maintenance_type} for asset ID {maintenance.asset_id}')
        flash('Maintenance scheduled successfully!', 'success')
        return redirect(url_for('view_maintenance'))

    assets = Asset.query.all()
    return render_template('add_maintenance.html', assets=assets)

@app.route('/maintenance/<int:maintenance_id>/complete', methods=['POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def complete_maintenance(maintenance_id):
    maintenance = AssetMaintenance.query.get_or_404(maintenance_id)

    maintenance.status = 'Completed'
    maintenance.completed_date = datetime.utcnow().date()
    maintenance.cost = float(request.form.get('cost', 0)) if request.form.get('cost') else None
    maintenance.maintenance_notes = request.form.get('notes', '')
    maintenance.updated_by = session['user_id']

    db.session.commit()

    log_activity(session['user_id'], 'Maintenance Completed', 
                f'Completed maintenance for asset {maintenance.asset.asset_tag}')
    flash('Maintenance marked as completed!', 'success')
    return redirect(url_for('asset_lifecycle_dashboard'))

@app.route('/view_depreciation')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def view_depreciation():
    depreciation_records = AssetDepreciation.query.order_by(AssetDepreciation.created_at.desc()).all()
    return render_template('depreciation.html', depreciation_records=depreciation_records)

@app.route('/custom_reports')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def custom_reports():
    return render_template('custom_reports.html')

# Analytics Dashboard
@app.route('/analytics')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def analytics_dashboard():
    user = User.query.get(session['user_id'])

    # Get basic counts
    total_requests = AssetRequest.query.count()
    approved_requests = AssetRequest.query.filter_by(status='Approved').count()
    pending_requests = AssetRequest.query.filter_by(status='Pending').count()
    fulfilled_requests = AssetRequest.query.filter_by(status='Fulfilled').count()
    rejected_requests = AssetRequest.query.filter_by(status='Rejected').count()

    total_assets = Asset.query.count()
    fixed_assets = Asset.query.filter_by(asset_type='Fixed Asset').count()
    consumable_assets = Asset.query.filter_by(asset_type='Consumable Asset').count()
    available_assets = Asset.query.filter_by(status='Available').count()
    in_use_assets = Asset.query.filter_by(status='In Use').count()

    total_vendors = Vendor.query.filter_by(is_active=True).count()
    total_quotations = ProcurementQuotation.query.count()

    total_maintenance = AssetMaintenance.query.count()
    completed_maintenance = AssetMaintenance.query.filter_by(status='Completed').count()
    pending_maintenance = AssetMaintenance.query.filter_by(status='Scheduled').count()

    # Calculate rates
    approval_rate = (approved_requests / total_requests * 100) if total_requests > 0 else 0
    utilization_rate = (in_use_assets / total_assets * 100) if total_assets > 0 else 0
    completion_rate = (completed_maintenance / total_maintenance * 100) if total_maintenance > 0 else 0
    quote_approval_rate = 75.0  # Placeholder

    # Calculate financial data
    total_estimated = sum([r.estimated_cost for r in AssetRequest.query.all() if r.estimated_cost])
    total_asset_value = sum([a.current_value for a in Asset.query.all() if a.current_value])
    total_bills = sum([b.bill_amount for b in Bill.query.all()])
    cost_variance = total_estimated - total_bills

    # Get department distribution
    dept_query = db.session.query(User.department, db.func.count(AssetRequest.id)).join(AssetRequest, User.id == AssetRequest.user_id).group_by(User.department).all()
    department_distribution = [(dept, count) for dept, count in dept_query]

    # Get monthly trends (last 6 months)
    from dateutil.relativedelta import relativedelta
    monthly_trends = []
    for i in range(6):
        month_start = datetime.utcnow() - relativedelta(months=i)
        month_end = month_start + relativedelta(months=1)
        count = AssetRequest.query.filter(
            AssetRequest.created_at >= month_start,
            AssetRequest.created_at < month_end
        ).count()
        monthly_trends.append({'month': month_start, 'count': count})
    monthly_trends.reverse()

    # Create analytics object
    class AnalyticsData:
        def __init__(self):
            self.request_stats = type('obj', (object,), {
                'total': total_requests,
                'pending': pending_requests,
                'approved': approved_requests,
                'fulfilled': fulfilled_requests,
                'approval_rate': approval_rate
            })()

            self.asset_stats = type('obj', (object,), {
                'total': total_assets,
                'fixed': fixed_assets,
                'consumable': consumable_assets,
                'available': available_assets,
                'in_use': in_use_assets,
                'utilization_rate': utilization_rate
            })()

            self.maintenance_stats = type('obj', (object,), {
                'total': total_maintenance,
                'completed': completed_maintenance,
                'pending': pending_maintenance,
                'completion_rate': completion_rate
            })()

            self.cost_stats = type('obj', (object,), {
                'total_estimated': total_estimated,
                'total_asset_value': total_asset_value,
                'total_bills': total_bills,
                'cost_variance': cost_variance
            })()

            self.vendor_stats = type('obj', (object,), {
                'total_vendors': total_vendors,
                'total_quotations': total_quotations,
                'quote_approval_rate': quote_approval_rate
            })()

            self.department_distribution = department_distribution
            self.monthly_trends = monthly_trends

    analytics = AnalyticsData()

    return render_template('analytics.html', 
                         user=user,
                         analytics=analytics)

# Quotation Management Routes
@app.route('/quotations')
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def view_quotations():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')

    query = ProcurementQuotation.query
    if status:
        query = query.filter_by(status=status)

    quotations = query.order_by(ProcurementQuotation.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False)

    user = User.query.get(session['user_id'])
    return render_template('quotations.html', 
                         quotations=quotations,
                         selected_status=status,
                         user=user)

@app.route('/quotation/add', methods=['GET', 'POST'])
@require_role(['Admin', 'MD', 'Accounts/SCM'])
def add_quotation():
    if request.method == 'POST':
        quotation = ProcurementQuotation()
        quotation.request_id = int(request.form['request_id']) if request.form['request_id'] else None
        quotation.vendor_id = int(request.form['vendor_id'])
        quotation.quotation_number = request.form['quotation_number']
        quotation.quoted_price = float(request.form['quoted_price'])
        quotation.quoted_quantity = int(request.form['quoted_quantity'])
        quotation.delivery_timeline = request.form.get('delivery_timeline', '')
        quotation.validity_period = int(request.form.get('validity_period', 30))
        quotation.payment_terms = request.form.get('payment_terms', '')
        quotation.warranty_period = request.form.get('warranty_period', '')
        quotation.specifications = request.form.get('specifications', '')
        quotation.additional_costs = float(request.form.get('additional_costs', 0))
        quotation.total_cost = quotation.quoted_price * quotation.quoted_quantity + quotation.additional_costs
        quotation.submitted_by = session['user_id']

        db.session.add(quotation)
        db.session.commit()

        log_activity(session['user_id'], 'Quotation Added', 
                    f'Added quotation {quotation.quotation_number} from {quotation.vendor.vendor_name}')
        flash('Quotation added successfully!', 'success')
        return redirect(url_for('view_quotations'))

    vendors = Vendor.query.filter_by(is_active=True).all()
    requests = AssetRequest.query.filter_by(status='Approved').all()
    return render_template('add_quotation.html', vendors=vendors, requests=requests)

# API Search Route for global search
@app.route('/api/search')
@require_login
def api_search():
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify({'results': []})

    user = User.query.get(session['user_id'])
    results = []

    # Search requests
    requests = AssetRequest.query.filter(
        AssetRequest.item_name.ilike(f'%{query}%')
    ).limit(5).all()

    for req in requests:
        results.append({
            'type': 'Request',
            'title': f"#{req.id} - {req.item_name}",
            'subtitle': f"Status: {req.status} | Requester: {req.requester.full_name}",
            'url': url_for('view_requests'),
            'icon': 'fas fa-list',
            'status': req.status
        })

    # Search assets
    assets = Asset.query.filter(
        or_(Asset.name.ilike(f'%{query}%'), Asset.asset_tag.ilike(f'%{query}%'))
    ).limit(5).all()

    for asset in assets:
        results.append({
            'type': 'Asset',
            'title': f"{asset.asset_tag} - {asset.name}",
            'subtitle': f"Category: {asset.category} | Status: {asset.status}",
            'url': url_for('view_asset_detail', asset_id=asset.id),
            'icon': 'fas fa-boxes',
            'status': asset.status
        })

    # Search vendors (if user has permission)
    if user.role in ['Admin', 'MD', 'Accounts/SCM']:
        vendors = Vendor.query.filter(
            Vendor.vendor_name.ilike(f'%{query}%')
        ).limit(5).all()

        for vendor in vendors:
            results.append({
                'type': 'Vendor',
                'title': vendor.vendor_name,
                'subtitle': f"Category: {vendor.category} | Contact: {vendor.contact_person}",
                'url': url_for('view_vendor_detail', vendor_id=vendor.id),
                'icon': 'fas fa-truck',
                'status': 'Active' if vendor.is_active else 'Inactive'
            })

    return jsonify({'results': results})

@app.route('/escalate-to-md/<int:request_id>', methods=['POST'])
@require_role(['Admin', 'Accounts/SCM'])
def escalate_to_md(request_id):
    """Escalate request directly to MD when there are issues"""
    user = User.query.get(session['user_id'])
    asset_request = AssetRequest.query.get_or_404(request_id)

    if asset_request.status not in ['Pending', 'Approved']:
        flash('Only pending or approved requests can be escalated to MD.', 'warning')
        return redirect(url_for('view_requests'))

    escalation_reason = request.form.get('escalation_reason', '')

    # Create escalation approval record
    approval = Approval()
    approval.request_id = request_id
    approval.approver_id = user.id
    approval.approval_level = asset_request.current_approval_level
    approval.action = 'Escalated to MD'
    approval.comments = f'Escalated to MD. Reason: {escalation_reason}'
    db.session.add(approval)

    # Update request to MD level
    asset_request.current_approval_level = 999  # Special MD level
    asset_request.status = 'Pending'
    asset_request.updated_at = datetime.utcnow()

    db.session.commit()

    log_activity(user.id, 'Request Escalated', 
                f'Escalated request #{request_id} to MD. Reason: {escalation_reason}',
                request_id)

    flash(f'Request #{request_id} has been escalated to MD for review.', 'info')
    return redirect(url_for('view_requests'))

@app.route('/delete-all-data')
@require_role(['MD'])
def delete_all_data():
    """Delete all application data - MD ONLY"""
    try:
        # Clear all data from tables
        db.session.execute(text('DELETE FROM activity_log'))
        db.session.execute(text('DELETE FROM approval'))
        db.session.execute(text('DELETE FROM uploaded_file'))
        db.session.execute(text('DELETE FROM asset_request'))
        db.session.execute(text('DELETE FROM bill'))
        db.session.execute(text('DELETE FROM inventory_update'))
        db.session.execute(text('DELETE FROM item_assignment'))
        db.session.execute(text('DELETE FROM asset_maintenance'))
        db.session.execute(text('DELETE FROM asset_depreciation'))
        db.session.execute(text('DELETE FROM warranty_alert'))
        db.session.execute(text('DELETE FROM procurement_quotation'))
        db.session.execute(text('DELETE FROM asset'))
        db.session.execute(text('DELETE FROM vendor'))
        db.session.execute(text('DELETE FROM user'))
        
        # Clear uploads folder
        uploads_folder = app.config['UPLOAD_FOLDER']
        if os.path.exists(uploads_folder):
            for filename in os.listdir(uploads_folder):
                if filename != '.gitkeep':
                    file_path = os.path.join(uploads_folder, filename)
                    try:
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                        elif os.path.isdir(file_path):
                            shutil.rmtree(file_path)
                    except Exception as e:
                        pass
        
        # Create new admin user
        admin_user = User()
        admin_user.username = 'admin'
        admin_user.email = 'admin@hexamed.com'
        admin_user.full_name = 'System Administrator'
        admin_user.role = 'MD'
        admin_user.floor = 'All'
        admin_user.department = 'Admin Block'
        admin_user.set_password('hexamed123')
        db.session.add(admin_user)
        
        # Create accounts user
        accounts_user = User()
        accounts_user.username = 'accounts'
        accounts_user.email = 'accounts@hexamed.com'
        accounts_user.full_name = 'Accounts Department'
        accounts_user.role = 'Accounts/SCM'
        accounts_user.floor = 'All'
        accounts_user.department = 'Accounts'
        accounts_user.set_password('accounts123')
        db.session.add(accounts_user)
        
        db.session.commit()
        
        # Log out current user
        session.clear()
        
        flash('All application data has been deleted successfully! Please login with admin/hexamed123', 'success')
        return redirect(url_for('login'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting data: {str(e)}', 'danger')
        return redirect(url_for('admin_panel'))

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('base.html', error='Internal server error'), 500